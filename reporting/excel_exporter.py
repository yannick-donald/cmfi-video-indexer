from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import UTC, datetime
from itertools import islice
from pathlib import Path
from typing import Any, BinaryIO, Iterable, Iterator

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from database.models import VideoRecord
from database.repository import SearchFilters, VideoRepository
from reporting.duplicates import find_duplicate_groups
from utils.formatters import format_bytes, format_duration

LOGGER = logging.getLogger(__name__)

SHEET_INVENTORY = "Inventaire"
SHEET_STATISTICS = "Statistiques"
SHEET_DUPLICATES = "Doublons"
SHEET_FOLDERS = "Arborescence"
SHEET_ERRORS = "Erreurs"

WORKFLOW_LABELS = {
    "digitized": "Numérisée",
    "to_review": "À visionner",
    "transcribed": "Transcrite",
    "treated": "Traitée",
    "ready_edit": "Prête montage",
    "published": "Publiée",
}

ASSET_LABELS = {"raw": "Brute", "cut": "Découpée"}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13, color="1F3864")
_HEADER_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

_INT_FORMAT = "#,##0"
_DECIMAL_FORMAT = "#,##0.00"

# Rows are pulled from SQLite in chunks of this size so that label lookups are
# batched instead of issuing one query per video.
_CHUNK_SIZE = 500


@dataclass(slots=True)
class ExportSummary:
    output_path: Path | None
    videos: int
    duplicate_groups: int
    duplicate_files: int
    folders: int
    errors: int


@dataclass(slots=True)
class _Column:
    header: str
    width: int
    number_format: str | None = None


_INVENTORY_COLUMNS: list[_Column] = [
    _Column("ID interne", 16),
    _Column("Nom du fichier", 46),
    _Column("Titre éditorial", 38),
    _Column("Dossier", 42),
    _Column("Lecteur partagé", 20),
    _Column("Type d'actif", 12),
    _Column("Étape de production", 18),
    _Column("Labels", 26),
    _Column("Responsable", 26),
    _Column("Complétude", 12, _INT_FORMAT),
    _Column("Intervenant", 22),
    _Column("Prédicateur", 22),
    _Column("Thème principal", 28),
    _Column("Type de contenu", 18),
    _Column("Événement", 24),
    _Column("Date événement", 14),
    _Column("Série", 22),
    _Column("Langue", 10),
    _Column("Taille (octets)", 16, _INT_FORMAT),
    _Column("Taille", 12),
    _Column("Durée (s)", 12, _DECIMAL_FORMAT),
    _Column("Durée", 12),
    _Column("Résolution", 12),
    _Column("FPS", 8, _DECIMAL_FORMAT),
    _Column("Codec vidéo", 12),
    _Column("Codec audio", 12),
    _Column("Débit (bps)", 14, _INT_FORMAT),
    _Column("Audio", 8),
    _Column("Extension", 10),
    _Column("Propriétaire", 26),
    _Column("Créé le", 20),
    _Column("Modifié le", 20),
    _Column("Première détection", 20),
    _Column("Revu le", 20),
    _Column("Statut scan", 12),
    _Column("Lien Drive", 42),
]

# The five completeness checks mirrored from the dashboard, so the report and
# the web UI never disagree on what "complete" means.
_COMPLETENESS_TOTAL = 5


def _completeness_score(video: VideoRecord, labels: list[dict[str, Any]]) -> int:
    return sum(
        (
            bool(video.reviewed_at),
            bool(video.editorial_title.strip()),
            bool(video.main_theme.strip()),
            bool((video.speaker or video.preacher).strip()),
            bool(labels),
        )
    )


def _chunked(iterable: Iterable[VideoRecord], size: int) -> Iterator[list[VideoRecord]]:
    iterator = iter(iterable)
    while chunk := list(islice(iterator, size)):
        yield chunk


def _style_header(sheet: Worksheet, columns: list[_Column], row: int = 1) -> None:
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=column.header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        sheet.column_dimensions[get_column_letter(index)].width = column.width
    sheet.row_dimensions[row].height = 28


def _finalize_table(sheet: Worksheet, columns: list[_Column], header_row: int = 1) -> None:
    """Freeze the header and add filter dropdowns over the populated range."""
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    last_row = max(sheet.max_row, header_row)
    last_column = get_column_letter(len(columns))
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"


def _apply_number_formats(sheet: Worksheet, columns: list[_Column], header_row: int = 1) -> None:
    for index, column in enumerate(columns, start=1):
        if not column.number_format:
            continue
        letter = get_column_letter(index)
        for row in range(header_row + 1, sheet.max_row + 1):
            sheet[f"{letter}{row}"].number_format = column.number_format


def _write_inventory(
    sheet: Worksheet,
    repo: VideoRepository,
    filters: SearchFilters | None,
) -> int:
    _style_header(sheet, _INVENTORY_COLUMNS)
    count = 0

    display_names = repo.get_user_display_names()

    for chunk in _chunked(repo.iter_videos(filters), _CHUNK_SIZE):
        labels_map = repo.get_labels_map([video.file_id for video in chunk])
        for video in chunk:
            labels = labels_map.get(video.file_id, [])
            sheet.append(
                [
                    video.internal_video_id,
                    video.file_name,
                    video.editorial_title,
                    video.folder_path,
                    video.shared_drive_name,
                    ASSET_LABELS.get(video.asset_type, video.asset_type),
                    WORKFLOW_LABELS.get(video.workflow_stage, video.workflow_stage),
                    ", ".join(str(label["name"]) for label in labels),
                    display_names.get(video.assigned_user_email, video.assigned_user_email),
                    f"{_completeness_score(video, labels)}/{_COMPLETENESS_TOTAL}",
                    video.speaker,
                    video.preacher,
                    video.main_theme,
                    video.content_type,
                    video.event_name,
                    video.event_date,
                    video.series_name,
                    video.language,
                    video.file_size,
                    format_bytes(video.file_size),
                    video.duration_seconds,
                    format_duration(video.duration_seconds),
                    video.resolution,
                    video.fps,
                    video.video_codec,
                    video.audio_codec,
                    video.bitrate,
                    "Oui" if video.has_audio else ("Non" if video.has_audio is False else ""),
                    video.file_extension,
                    video.owner,
                    video.created_at,
                    video.modified_at,
                    video.first_seen_at,
                    video.reviewed_at,
                    video.scan_status,
                    video.drive_url,
                ]
            )
            count += 1

    _apply_number_formats(sheet, _INVENTORY_COLUMNS)
    _finalize_table(sheet, _INVENTORY_COLUMNS)
    return count


def _write_section(sheet: Worksheet, row: int, title: str) -> int:
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT
    return row + 1


def _write_statistics(
    sheet: Worksheet,
    repo: VideoRepository,
    *,
    generated_at: str,
    duplicate_groups: int,
    duplicate_files: int,
    folder_count: int,
    error_count: int,
) -> None:
    stats = repo.get_stats()
    workflow = repo.get_workflow_stats()
    tracking = repo.get_tracking_stats()

    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 18

    row = _write_section(sheet, 1, "Vue d'ensemble")
    overview = [
        ("Rapport généré le", generated_at),
        ("Vidéos indexées", stats["total_videos"]),
        ("Volume total", format_bytes(stats["total_bytes"])),
        ("Volume total (octets)", stats["total_bytes"]),
        ("Durée totale", format_duration(stats["total_duration_seconds"])),
        ("Dossiers distincts", folder_count),
        ("Groupes de doublons", duplicate_groups),
        ("Fichiers en doublon", duplicate_files),
        ("Éléments en erreur", error_count),
    ]
    for name, value in overview:
        sheet.cell(row=row, column=1, value=name).font = Font(bold=True)
        cell = sheet.cell(row=row, column=2, value=value)
        if isinstance(value, int):
            cell.number_format = _INT_FORMAT
        row += 1

    row += 1
    row = _write_section(sheet, row, "Étapes de production")
    stages = workflow.get("stages", {})
    for key, label in WORKFLOW_LABELS.items():
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=int(stages.get(key, 0))).number_format = _INT_FORMAT
        row += 1

    row += 1
    row = _write_section(sheet, row, "Types d'actif")
    asset_types = workflow.get("assets", {})
    for key, label in ASSET_LABELS.items():
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=int(asset_types.get(key, 0))).number_format = _INT_FORMAT
        row += 1

    row += 1
    row = _write_section(sheet, row, "Suivi éditorial")
    tracking_labels = {
        "new_count": "Nouvelles vidéos",
        "incomplete": "Fiches à compléter",
        "missing_labels": "Sans label",
        "unlinked_cuts": "Découpées sans source",
    }
    for key, label in tracking_labels.items():
        if key not in tracking:
            continue
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=int(tracking[key])).number_format = _INT_FORMAT
        row += 1

    row += 1
    row = _write_section(sheet, row, "Formats les plus fréquents")
    for entry in stats["top_formats"]:
        sheet.cell(row=row, column=1, value=entry["label"])
        sheet.cell(row=row, column=2, value=entry["count"]).number_format = _INT_FORMAT
        row += 1

    row += 1
    row = _write_section(sheet, row, "Résolutions les plus fréquentes")
    for entry in stats["top_resolutions"]:
        sheet.cell(row=row, column=1, value=entry["label"])
        sheet.cell(row=row, column=2, value=entry["count"]).number_format = _INT_FORMAT
        row += 1

    row += 1
    row = _write_section(sheet, row, "Fichiers les plus volumineux")
    for entry in stats["largest_files"]:
        sheet.cell(row=row, column=1, value=entry["file_name"])
        sheet.cell(row=row, column=2, value=format_bytes(entry["file_size"]))
        sheet.cell(row=row, column=3, value=entry["folder_path"])
        row += 1

    row += 1
    row = _write_section(sheet, row, "Vidéos les plus longues")
    for entry in stats["longest_videos"]:
        sheet.cell(row=row, column=1, value=entry["file_name"])
        sheet.cell(row=row, column=2, value=format_duration(entry["duration_seconds"]))
        sheet.cell(row=row, column=3, value=entry["folder_path"])
        row += 1

    sheet.freeze_panes = "A2"


_DUPLICATE_COLUMNS = [
    _Column("Groupe", 10, _INT_FORMAT),
    _Column("Similarité", 14),
    _Column("Motif du rapprochement", 40),
    _Column("Fichiers", 10, _INT_FORMAT),
    _Column("Espace récupérable", 18),
    _Column("Nom du fichier", 46),
    _Column("Dossier", 42),
    _Column("ID interne", 16),
    _Column("Taille (octets)", 16, _INT_FORMAT),
    _Column("Durée (s)", 12, _DECIMAL_FORMAT),
    _Column("Modifié le", 20),
    _Column("Lien Drive", 42),
]


def _write_duplicates(sheet: Worksheet, groups: list[Any]) -> tuple[int, int]:
    _style_header(sheet, _DUPLICATE_COLUMNS)
    files = 0
    for index, group in enumerate(groups, start=1):
        for position, member in enumerate(group.members):
            # Group-level columns are written once, on the group's first row, so
            # the sheet stays readable when sorted or filtered by group.
            first = position == 0
            sheet.append(
                [
                    index if first else None,
                    group.confidence if first else None,
                    group.reason if first else None,
                    len(group.members) if first else None,
                    format_bytes(group.wasted_bytes) if first else None,
                    member.file_name,
                    member.folder_path,
                    member.internal_video_id,
                    member.file_size,
                    member.duration_seconds,
                    member.modified_at,
                    member.drive_url,
                ]
            )
            files += 1
    _apply_number_formats(sheet, _DUPLICATE_COLUMNS)
    _finalize_table(sheet, _DUPLICATE_COLUMNS)
    return len(groups), files


_FOLDER_COLUMNS = [
    _Column("Dossier", 52),
    _Column("Dossier parent", 28),
    _Column("Lecteur partagé", 20),
    _Column("Profondeur", 12, _INT_FORMAT),
    _Column("Vidéos", 10, _INT_FORMAT),
    _Column("Brutes", 10, _INT_FORMAT),
    _Column("Découpées", 12, _INT_FORMAT),
    _Column("Volume", 14),
    _Column("Volume (octets)", 18, _INT_FORMAT),
    _Column("Durée totale", 14),
    _Column("Modifié du", 20),
    _Column("Modifié au", 20),
]


def _write_folders(sheet: Worksheet, folders: list[dict[str, Any]]) -> int:
    _style_header(sheet, _FOLDER_COLUMNS)
    for folder in folders:
        path = str(folder["folder_path"])
        depth = len([segment for segment in path.split("/") if segment])
        sheet.append(
            [
                path,
                folder["parent_folder"],
                folder["shared_drive_name"],
                depth,
                folder["video_count"],
                folder["raw_count"],
                folder["cut_count"],
                format_bytes(folder["total_bytes"]),
                folder["total_bytes"],
                format_duration(folder["total_duration_seconds"]),
                folder["first_modified_at"] or "",
                folder["last_modified_at"] or "",
            ]
        )
    _apply_number_formats(sheet, _FOLDER_COLUMNS)
    _finalize_table(sheet, _FOLDER_COLUMNS)
    return len(folders)


_ERROR_COLUMNS = [
    _Column("Origine", 16),
    _Column("Nom du fichier", 46),
    _Column("Dossier", 42),
    _Column("Statut", 14),
    _Column("Message", 62),
    _Column("Horodatage", 22),
    _Column("Lien Drive", 42),
]


def _write_errors(
    sheet: Worksheet,
    error_records: list[dict[str, Any]],
    scan_runs: list[dict[str, Any]],
) -> int:
    _style_header(sheet, _ERROR_COLUMNS)
    for record in error_records:
        sheet.append(
            [
                "Fichier",
                record["file_name"],
                record["folder_path"],
                record["scan_status"] or "",
                record["error_message"] or "",
                record["last_scanned_at"] or "",
                record["drive_url"] or "",
            ]
        )
    for run in scan_runs:
        if not run["errors"] and run["status"] == "succeeded":
            continue
        sheet.append(
            [
                "Scan",
                f"Scan du {run['started_at']}",
                "",
                run["status"] or "",
                f"{run['errors'] or 0} erreur(s), {run['videos_found'] or 0} vidéo(s) trouvée(s)",
                run["finished_at"] or run["started_at"] or "",
                "",
            ]
        )
    _finalize_table(sheet, _ERROR_COLUMNS)
    return len(error_records)


def build_workbook(
    repo: VideoRepository,
    filters: SearchFilters | None = None,
) -> tuple[Workbook, ExportSummary]:
    """Build the full multi-sheet inventory workbook in memory."""
    generated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")

    workbook = Workbook()
    inventory_sheet = workbook.active
    inventory_sheet.title = SHEET_INVENTORY
    statistics_sheet = workbook.create_sheet(SHEET_STATISTICS)
    duplicates_sheet = workbook.create_sheet(SHEET_DUPLICATES)
    folders_sheet = workbook.create_sheet(SHEET_FOLDERS)
    errors_sheet = workbook.create_sheet(SHEET_ERRORS)

    video_count = _write_inventory(inventory_sheet, repo, filters)

    groups = find_duplicate_groups(repo.get_duplicate_candidates())
    duplicate_groups, duplicate_files = _write_duplicates(duplicates_sheet, groups)

    folders = repo.get_folder_summary()
    folder_count = _write_folders(folders_sheet, folders)

    error_records = repo.get_error_records()
    error_count = _write_errors(errors_sheet, error_records, repo.get_recent_scan_runs())

    _write_statistics(
        statistics_sheet,
        repo,
        generated_at=generated_at,
        duplicate_groups=duplicate_groups,
        duplicate_files=duplicate_files,
        folder_count=folder_count,
        error_count=error_count,
    )

    summary = ExportSummary(
        output_path=None,
        videos=video_count,
        duplicate_groups=duplicate_groups,
        duplicate_files=duplicate_files,
        folders=folder_count,
        errors=error_count,
    )
    return workbook, summary


def export_to_stream(
    repo: VideoRepository,
    stream: BinaryIO,
    filters: SearchFilters | None = None,
) -> ExportSummary:
    workbook, summary = build_workbook(repo, filters)
    workbook.save(stream)
    return summary


def export_to_path(
    repo: VideoRepository,
    output_path: Path,
    filters: SearchFilters | None = None,
) -> ExportSummary:
    workbook, summary = build_workbook(repo, filters)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    summary.output_path = output_path
    LOGGER.info(
        "Excel inventory written to %s (%s videos, %s duplicate groups, %s folders, %s errors)",
        output_path,
        summary.videos,
        summary.duplicate_groups,
        summary.folders,
        summary.errors,
    )
    return summary
